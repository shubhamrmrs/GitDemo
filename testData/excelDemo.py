import openpyxl

book = openpyxl.load_workbook("D:\Software\IMP\Projects\pythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)  # reading data

print(sheet.max_row)
print(sheet.max_column)

# sheet.cell(row=2, column=2).value = "Rahul"      # writing data
# print(sheet.cell(row=2, column=2).value)
print(sheet['A3'].value)

print("************************************************")
# for i in range(2, sheet.max_row + 1):    # to get rows
#     for j in range(1, sheet.max_column + 1):     # to get columns
#         print(sheet.cell(row=i, column=j).value)

Dict = {}

print("************************************************")
for i in range(2, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "4th_Test":
        for j in range(2, sheet.max_column + 1):
            # print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
