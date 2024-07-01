import openpyxl


class HomePageData:
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}

        book = openpyxl.load_workbook("D:\Software\IMP\Projects\pythonDemo.xlsx")
        sheet = book.active
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value in test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
